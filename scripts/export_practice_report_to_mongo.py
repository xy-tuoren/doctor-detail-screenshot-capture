r"""将「医生执业医院信息」xlsx 导入 MongoDB（仅源数据 + 互联网医院映射，派生数据不落库）。

用法（PowerShell）：

    $env:MONGO_URI = "mongodb://hospital_admin_user:***@43.138.254.94:27017/hospital_admin?authSource=hospital_admin"
    & .venv\Scripts\python.exe scripts\export_practice_report_to_mongo.ps1 `
        -XlsxPath "workspace/artifacts/医生执业医院信息_20260706.xlsx"

Collection 设计（数据库 hospital_admin）：

1. practiceDetails   全量明细源数据（每行 = 每医生×每医院），按 reportId 绑定导入批次
2. internetHospitals 互联网医院名单映射（主数据，upsert，方便后续新增）
3. importReports     导入批次元数据（1 条/次，含汇总统计）

派生数据（医生执业医院数 / 医院重叠数 / 互联网执业医生数 / 分桶图）不落库，
查询时用 aggregation pipeline 现算。
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymongo import ASCENDING, MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

# 全量明细 20 列（与 src/minke_reg/practice_table.py DETAIL_HEADERS + 3 莲藕列一致）
DETAIL_HEADERS: tuple[str, ...] = (
    "姓名", "身份证号", "执业证书编码", "性别",
    "医师类别", "医师级别", "执业范围", "任职资格",
    "审批日期", "开始日期", "结束日期",
    "是否主执业机构", "是否省外",
    "执业医院", "医院地址", "省份", "数据来源",
    "档案编号", "档案状态", "所属团队",
)

# 互联网医院映射列（不存「执业医生数」，查询时聚合算）
INTERNET_HEADERS: tuple[str, ...] = (
    "互联网医院名单", "经营现状", "网院属性", "医院牌照类型",
    "网院公司名称", "实体医院", "入驻平台",
)

# 日期列 → ISODate；空字符串/None → None
DATE_FIELDS: dict[str, str] = {
    "审批日期": "auditDate",
    "开始日期": "startDate",
    "结束日期": "endDate",
}

# 布尔列：是/否 → True/False
BOOL_FIELDS: dict[str, str] = {
    "是否主执业机构": "isMainPractice",
    "是否省外": "isOutOfProvince",
}


def normalize_cert(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "")


def parse_date(value: Any) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip()
    # 兼容 2020-09-15 / 2020/9/15 / datetime 对象
    if isinstance(value, datetime):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None  # 解析失败保留 None，不阻断导入


def yes_no_to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    s = str(value).strip()
    if s == "是":
        return True
    if s == "否":
        return False
    return None


def read_sheet_rows(path: Path, sheet_name: str, headers: tuple[str, ...]) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"xlsx 缺少 sheet「{sheet_name}」")
    ws = wb[sheet_name]
    col_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(col_headers) if h}
    missing = [h for h in headers if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"sheet「{sheet_name}」缺少列: {missing}")
    rows: list[dict[str, Any]] = []
    for line in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in line):
            continue
        row = {h: line[idx[h]] for h in headers}
        rows.append(row)
    wb.close()
    return rows


def detail_row_to_doc(row: dict[str, Any], report_id, imported_at: datetime) -> dict[str, Any]:
    doc: dict[str, Any] = {
        "reportId": report_id,
        "importedAt": imported_at,
        "doctorName": str(row.get("姓名") or "").strip(),
        "idCard": str(row.get("身份证号") or "").strip(),
        "practicingCertCode": normalize_cert(row.get("执业证书编码")),
        "gender": str(row.get("性别") or "").strip(),
        "medicalCategory": str(row.get("医师类别") or "").strip(),
        "medicalLevel": str(row.get("医师级别") or "").strip(),
        "practiceScope": str(row.get("执业范围") or "").strip(),
        "qualificationTitle": str(row.get("任职资格") or "").strip(),
        "isMainPractice": yes_no_to_bool(row.get("是否主执业机构")),
        "isOutOfProvince": yes_no_to_bool(row.get("是否省外")),
        "practiceHospital": str(row.get("执业医院") or "").strip(),
        "hospitalAddress": str(row.get("医院地址") or "").strip(),
        "province": str(row.get("省份") or "").strip(),
        "dataSource": str(row.get("数据来源") or "").strip(),
        "doctorFileId": str(row.get("档案编号") or "").strip(),
        "archiveStatus": str(row.get("档案状态") or "").strip(),
        "team": str(row.get("所属团队") or "").strip(),
    }
    for cn, en in DATE_FIELDS.items():
        doc[en] = parse_date(row.get(cn))
    return doc


def internet_row_to_doc(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "internetHospitalName": str(row.get("互联网医院名单") or "").strip(),
        "operationStatus": str(row.get("经营现状") or "").strip(),
        "hospitalAttribute": str(row.get("网院属性") or "").strip(),
        "licenseType": str(row.get("医院牌照类型") or "").strip(),
        "networkCompany": str(row.get("网院公司名称") or "").strip(),
        "entityHospital": str(row.get("实体医院") or "").strip(),
        "platform": str(row.get("入驻平台") or "").strip(),
    }


def ensure_indexes(db) -> None:
    db.practiceDetails.create_index([("reportId", ASCENDING), ("practicingCertCode", ASCENDING)])
    db.practiceDetails.create_index([("reportId", ASCENDING), ("practiceHospital", ASCENDING)])
    db.practiceDetails.create_index([("reportId", ASCENDING), ("doctorName", ASCENDING)])
    db.practiceDetails.create_index([("reportId", ASCENDING), ("archiveStatus", ASCENDING)])
    db.practiceDetails.create_index([("reportId", ASCENDING), ("practiceHospital", ASCENDING), ("archiveStatus", ASCENDING)])
    # 互联网医院主数据：名单唯一（方便新增/去重），实体医院索引用于 join
    db.internetHospitals.create_index("internetHospitalName", unique=True)
    db.internetHospitals.create_index("entityHospital")
    db.internetHospitals.create_index([("active", ASCENDING), ("entityHospital", ASCENDING)])
    db.importReports.create_index([("importedAt", ASCENDING)])


def batch_insert(collection, docs: list[dict[str, Any]], batch_size: int = 1000) -> int:
    inserted = 0
    for i in range(0, len(docs), batch_size):
        chunk = docs[i:i + batch_size]
        res = collection.insert_many(chunk, ordered=False)
        inserted += len(res.inserted_ids)
    return inserted


def upsert_internet_hospitals(collection, docs: list[dict[str, Any]]) -> tuple[int, int]:
    """以 internetHospitalName 为唯一键 upsert：已有更新，新名单插入。返回 (matched, upserted)。"""
    now = datetime.now(timezone.utc)
    ops: list[UpdateOne] = []
    for d in docs:
        name = d["internetHospitalName"]
        if not name:
            continue
        ops.append(UpdateOne(
            {"internetHospitalName": name},
            {
                "$set": {
                    **d,
                    "active": True,
                    "updatedAt": now,
                },
                "$setOnInsert": {"createdAt": now},
            },
            upsert=True,
        ))
    if not ops:
        return (0, 0)
    res = collection.bulk_write(ops, ordered=False)
    return (res.matched_count, res.upserted_count)


def _load_dotenv(path: Path) -> None:
    """从项目根 .env 加载未设置的环境变量（不覆盖已有 os.environ）。"""
    if not path.is_file():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def main() -> int:
    parser = argparse.ArgumentParser(description="导入「医生执业医院信息」xlsx 到 MongoDB")
    parser.add_argument("-x", "--xlsx-path", required=True, help="xlsx 文件路径")
    parser.add_argument("--db-name", default="hospital_admin", help="数据库名（默认 hospital_admin）")
    parser.add_argument("--drop-existing", action="store_true",
                        help="导入前清空 practiceDetails / importReports（internetHospitals 主数据不动）")
    parser.add_argument("--dry-run", action="store_true", help="只解析不写库")
    args = parser.parse_args()

    _load_dotenv(Path(__file__).resolve().parent.parent / ".env")

    mongo_uri = os.environ.get("MONGO_URI")
    if not mongo_uri:
        print("[ERROR] 未设置环境变量 MONGO_URI（可写项目根 .env 或 $env:MONGO_URI）", file=sys.stderr)
        return 1

    xlsx_path = Path(args.xlsx_path).resolve()
    if not xlsx_path.exists():
        print(f"[ERROR] 文件不存在: {xlsx_path}", file=sys.stderr)
        return 1

    print(f"读取 xlsx: {xlsx_path}")
    detail_rows = read_sheet_rows(xlsx_path, "全量明细", DETAIL_HEADERS)
    internet_rows = read_sheet_rows(xlsx_path, "互联网医院重叠数", INTERNET_HEADERS)
    print(f"  全量明细: {len(detail_rows)} 行")
    print(f"  互联网医院映射: {len(internet_rows)} 行")

    if args.dry_run:
        print("[dry-run] 不写库，结束。")
        return 0

    print(f"连接 MongoDB: {args.db_name}")
    client = MongoClient(mongo_uri)
    db = client[args.db_name]

    if args.drop_existing:
        print("[!] 清空 practiceDetails / importReports")
        db.practiceDetails.delete_many({})
        db.importReports.delete_many({})

    print("创建索引...")
    ensure_indexes(db)

    imported_at = datetime.now(timezone.utc)
    report_doc = {
        "importedAt": imported_at,
        "sourceXlsx": str(xlsx_path),
        "status": "building",
    }
    insert_report_res = db.importReports.insert_one(report_doc)
    report_id = insert_report_res.inserted_id
    print(f"创建导入批次: reportId={report_id}")

    print("写入 practiceDetails...")
    detail_docs = [detail_row_to_doc(r, report_id, imported_at) for r in detail_rows]
    n_detail = batch_insert(db.practiceDetails, detail_docs)
    print(f"  已写入 {n_detail} 条")

    print("upsert internetHospitals（主数据，保留历史，新增新名单）...")
    internet_docs = [internet_row_to_doc(r) for r in internet_rows]
    matched, upserted = upsert_internet_hospitals(db.internetHospitals, internet_docs)
    total_inet = db.internetHospitals.count_documents({"active": True})
    print(f"  本批: 匹配已存在 {matched} / 新增 {upserted}; 主数据当前共 {total_inet} 条")

    # 汇总统计（不落派生数据，只落批次元信息）
    n_doctors = len({d["practicingCertCode"] for d in detail_docs if d["practicingCertCode"]})
    n_hospitals = len({d["practiceHospital"] for d in detail_docs if d["practiceHospital"]})
    n_matched = sum(1 for d in detail_docs if d["doctorFileId"])
    n_coop = len(db.practiceDetails.distinct("practicingCertCode", {"reportId": report_id, "archiveStatus": "启用"}))
    n_noncoop = len(db.practiceDetails.distinct("practicingCertCode", {"reportId": report_id, "archiveStatus": "停用"}))
    summary = {
        "detailRows": n_detail,
        "doctors": n_doctors,
        "lianouMatchedRows": n_matched,
        "lianouMissRows": n_detail - n_matched,
        "hospitals": n_hospitals,
        "coopDoctors": n_coop,
        "nonCoopDoctors": n_noncoop,
        "internetHospitals": total_inet,
    }
    db.importReports.update_one(
        {"_id": report_id},
        {"$set": {"status": "ready", "summary": summary}},
    )

    print("\n导入完成 ✓")
    print(f"  reportId:        {report_id}")
    print(f"  practiceDetails: {n_detail}")
    print(f"  doctors:         {n_doctors}")
    print(f"  hospitals:       {n_hospitals}")
    print(f"  internetHospitals(主数据): {total_inet} (本批新增 {upserted})")
    print(f"  合作/不合作:      {n_coop} / {n_noncoop}")
    print(f"\n后续查最新批次: db.importReports.find().sort({{importedAt:-1}}).limit(1)")
    client.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
