"""Parse institution UI/SOAP export files.

两种导出目录（均在 ``exports/`` 下）：
- ``exports/ui/``：机构端客户端 UI 手动导出（多为 ``.xls``，OOXML 伪装）
- ``exports/reg-api/``：SOAP 接口自动导出（标准 ``.xlsx``，openpyxl）

核对时在两目录中按文件名前缀各取修改时间最新的一份。
"""

from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .paths import EXPORT_SOURCE_DIRS

MAIN_HEADERS = frozenset({"审核日期"})
MULTI_HEADERS = frozenset({"开始日期", "结束日期"})

CERT_KEYS = ("执业证书编码", "执业证书编号")


@dataclass(frozen=True)
class ExportFiles:
    main: Path | None
    multi: Path | None


def _export_dirs(exports_dir: Path) -> list[Path]:
    """在 exports/ui、exports/reg-api 子目录查找；兼容仍放在 exports 根目录的旧文件。"""
    dirs: list[Path] = []
    for name in EXPORT_SOURCE_DIRS:
        sub = exports_dir / name
        if sub.is_dir():
            dirs.append(sub)
    if exports_dir.is_dir():
        dirs.append(exports_dir)
    return dirs


def _latest_export(exports_dir: Path, prefix: str) -> Path | None:
    candidates: list[Path] = []
    for directory in _export_dirs(exports_dir):
        for pattern in (f"{prefix}-*.xlsx", f"{prefix}-*.xls"):
            candidates.extend(directory.glob(pattern))
    candidates = [p for p in candidates if p.stat().st_size > 0]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_latest_exports(exports_dir: Path) -> ExportFiles:
    return ExportFiles(
        main=_latest_export(exports_dir, "主执业导出"),
        multi=_latest_export(exports_dir, "多执业导出"),
    )


def parse_export_file(path: Path) -> list[dict[str, Any]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    if path.suffix.lower() == ".xlsx":
        return _parse_xlsx(path)
    return _parse_legacy_xls(path)


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        records: list[dict[str, Any]] = []
        for row in row_iter:
            record: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                if value is None:
                    continue
                record[header] = value
            if record:
                records.append(record)
        return records
    finally:
        wb.close()


def _parse_legacy_xls(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as zf:
        shared_strings = _read_shared_strings(zf)
        rows = _read_sheet_rows(zf, shared_strings)

    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if value is None:
                continue
            record[header] = value
        if record:
            records.append(record)
    return records


def classify_export(headers: list[str]) -> str:
    header_set = set(headers)
    if MULTI_HEADERS.issubset(header_set):
        return "multi"
    if MAIN_HEADERS.intersection(header_set):
        return "main"
    raise ValueError(f"无法识别导出类型，表头: {headers}")


def normalize_cert_code(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def extract_cert_code(row: dict[str, Any]) -> str:
    for key in CERT_KEYS:
        value = row.get(key)
        if value is not None and str(value).strip():
            return normalize_cert_code(value)
    return ""


def build_export_index(exports_dir: Path) -> dict[str, Any]:
    """按执业证书编码建立索引（main: dict、multi: list）。

    身份证/姓名仍随行保留：姓名用于双字段校验，身份证供后续图片采集。
    """
    files = find_latest_exports(exports_dir)
    main_records = parse_export_file(files.main) if files.main else []
    multi_records = parse_export_file(files.multi) if files.multi else []

    main_index: dict[str, dict[str, Any]] = {}
    for row in main_records:
        cert = extract_cert_code(row)
        if cert:
            main_index[cert] = row

    multi_index: dict[str, list[dict[str, Any]]] = {}
    for row in multi_records:
        cert = extract_cert_code(row)
        if not cert:
            continue
        multi_index.setdefault(cert, []).append(row)

    return {
        "sources": {
            "main": str(files.main) if files.main else None,
            "multi": str(files.multi) if files.multi else None,
        },
        "counts": {
            "main": len(main_records),
            "multi": len(multi_records),
        },
        "main": main_index,
        "multi": multi_index,
    }


def _normalize_id(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip().upper())


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings: list[str] = []
    for si in root.findall(f"{ns}si"):
        strings.append("".join(t.text or "" for t in si.iter(f"{ns}t")))
    return strings


def _read_sheet_rows(zf: zipfile.ZipFile, shared_strings: list[str]) -> list[list[Any]]:
    root = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    sheet_data = root.find(f"{ns}sheetData")
    if sheet_data is None:
        return []

    rows: list[list[Any]] = []
    for row in sheet_data.findall(f"{ns}row"):
        cells: dict[int, Any] = {}
        for cell in row.findall(f"{ns}c"):
            ref = cell.get("r", "")
            col_idx = _column_index(ref)
            value_node = cell.find(f"{ns}v")
            if value_node is None:
                continue
            cell_type = cell.get("t")
            if cell_type == "s":
                cells[col_idx] = shared_strings[int(value_node.text)]
            else:
                cells[col_idx] = value_node.text
        if not cells:
            rows.append([])
            continue
        max_col = max(cells)
        rows.append([cells.get(i) for i in range(max_col + 1)])
    return rows


def _column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch.upper()) - ord("A") + 1)
    return index - 1
