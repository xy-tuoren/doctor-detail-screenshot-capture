"""从机构端执业明细 + 莲藕 GetDoctorMedicalPage 生成「医生执业医院信息」报告。

独立流程：不读 practice_hospitals_all_fixed，不读 doctors_api_cache（除非调用方传入缓存）。
"""

from __future__ import annotations

import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.minke_reg.practice_table import DETAIL_HEADERS

# 全量明细 20 列（机构端 17 列 + 莲藕 3 列；执业证书编码已在 DETAIL_HEADERS 第 3 列）
FULL_DETAIL_HEADERS: tuple[str, ...] = DETAIL_HEADERS + (
    "档案编号",
    "档案状态",
    "所属团队",
)

HOSPITAL_COUNT_BIN_ORDER: tuple[str, ...] = (
    "≥50",
    "49~40",
    "39~30",
    "29~20",
    "19~15",
    "14~10",
    "9~6",
    "5",
    "4",
    "3",
    "2",
    "1",
)

_CENTER = Alignment(horizontal="center", vertical="center")


def default_report_output_path(
    artifacts_dir: Path | str,
    *,
    on_date: date | None = None,
) -> Path:
    """默认报告路径：``医生执业医院信息_YYYYMMDD.xlsx``（按当天日期）。"""
    day = on_date or date.today()
    return Path(artifacts_dir) / f"医生执业医院信息_{day.strftime('%Y%m%d')}.xlsx"


def normalize_cert(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "")


def hospital_count_bin(count: int) -> str:
    if count >= 50:
        return "≥50"
    if count >= 40:
        return "49~40"
    if count >= 30:
        return "39~30"
    if count >= 20:
        return "29~20"
    if count >= 15:
        return "19~15"
    if count >= 10:
        return "14~10"
    if count >= 6:
        return "9~6"
    if count in (5, 4, 3, 2, 1):
        return str(count)
    return "1"


def _archive_status_label(staus: Any) -> str:
    if staus == 0:
        return "启用"
    if staus == 1:
        return "停用"
    if staus is None or staus == "":
        return ""
    return str(staus)


def build_lianou_indexes(
    records: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    by_cert: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in records:
        cert = normalize_cert(rec.get("practicingCertCode"))
        name = str(rec.get("doctorName") or "").strip()
        if cert and cert not in by_cert:
            by_cert[cert] = rec
        if name:
            by_name[name].append(rec)
    return by_cert, by_name


def lookup_lianou(
    by_cert: dict[str, dict[str, Any]],
    by_name: dict[str, list[dict[str, Any]]],
    *,
    cert: str,
    name: str,
) -> dict[str, Any] | None:
    norm = normalize_cert(cert)
    if norm and norm in by_cert:
        return by_cert[norm]
    hits = by_name.get(name.strip(), [])
    if len(hits) == 1:
        return hits[0]
    return None


def read_institution_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    index = {str(h): i for i, h in enumerate(headers) if h}
    missing = [h for h in DETAIL_HEADERS if h not in index]
    if missing:
        wb.close()
        raise ValueError(f"执业明细缺少列: {missing}（请用最新 fetch-practice-table 重新导出）")
    rows: list[dict[str, str]] = []
    for line in ws.iter_rows(min_row=2, values_only=True):
        if not any(line):
            continue
        row = {h: str(line[index[h]] or "") if line[index[h]] is not None else "" for h in DETAIL_HEADERS}
        rows.append(row)
    wb.close()
    return rows


def enrich_detail_rows(
    institution_rows: list[dict[str, str]],
    lianou_records: list[dict[str, Any]],
) -> list[dict[str, str]]:
    by_cert, by_name = build_lianou_indexes(lianou_records)
    out: list[dict[str, str]] = []
    for inst in institution_rows:
        cert = inst.get("执业证书编码", "")
        name = inst.get("姓名", "")
        lianou = lookup_lianou(by_cert, by_name, cert=cert, name=name)
        file_id = ""
        status = ""
        team = ""
        if lianou:
            file_id = str(lianou.get("doctorFileId") or "")
            status = _archive_status_label(lianou.get("staus"))
            team = str(lianou.get("teamName") or "")
        row = dict(inst)
        row["档案编号"] = file_id
        row["档案状态"] = status
        row["所属团队"] = team
        out.append(row)
    return out


@dataclass
class DoctorStat:
    file_id: str
    name: str
    team: str
    status: str
    cert: str
    hospitals: set[str] = field(default_factory=set)
    internet_hospitals: set[str] = field(default_factory=set)

    @property
    def hospital_count(self) -> int:
        return len(self.hospitals)

    @property
    def internet_count(self) -> int:
        return len(self.internet_hospitals)


def aggregate_doctors(
    detail_rows: list[dict[str, str]],
    internet_entities: set[str],
) -> dict[str, DoctorStat]:
    stats: dict[str, DoctorStat] = {}
    for row in detail_rows:
        cert = normalize_cert(row.get("执业证书编码"))
        name = row.get("姓名", "")
        key = cert or f"{name}|{row.get('身份证号', '')}"
        if key not in stats:
            stats[key] = DoctorStat(
                file_id=row.get("档案编号", ""),
                name=name,
                team=row.get("所属团队", ""),
                status=row.get("档案状态", ""),
                cert=cert,
            )
        hospital = str(row.get("执业医院") or "").strip()
        if hospital:
            stats[key].hospitals.add(hospital)
            if hospital in internet_entities:
                stats[key].internet_hospitals.add(hospital)
    return stats


def aggregate_hospital_overlap(detail_rows: list[dict[str, str]]) -> list[tuple[str, int, int, int]]:
    enabled: Counter[str] = Counter()
    disabled: Counter[str] = Counter()
    for row in detail_rows:
        hospital = str(row.get("执业医院") or "").strip()
        if not hospital:
            continue
        status = row.get("档案状态", "")
        if status == "启用":
            enabled[hospital] += 1
        elif status == "停用":
            disabled[hospital] += 1
    hospitals = sorted(set(enabled) | set(disabled))
    result: list[tuple[str, int, int, int]] = []
    for h in hospitals:
        en = enabled[h]
        dis = disabled[h]
        result.append((h, en, dis, en + dis))
    result.sort(key=lambda x: (-x[3], x[0]))
    return result


def read_internet_entities(template_path: Path) -> tuple[list[dict[str, Any]], set[str]]:
    wb = load_workbook(template_path, read_only=True, data_only=True)
    if "互联网医院重叠数" not in wb.sheetnames:
        wb.close()
        raise ValueError("模板缺少 sheet「互联网医院重叠数」")
    ws = wb["互联网医院重叠数"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i for i, h in enumerate(headers) if h}
    entity_i = idx.get("实体医院")
    if entity_i is None:
        wb.close()
        raise ValueError("模板「互联网医院重叠数」缺少「实体医院」列")
    rows: list[dict[str, Any]] = []
    entities: set[str] = set()
    for line in ws.iter_rows(min_row=2, values_only=True):
        if not any(line):
            continue
        entity = str(line[entity_i] or "").strip()
        row_dict = {
            str(headers[i]): line[i] if i < len(line) else ""
            for i in range(len(headers))
        }
        rows.append(row_dict)
        if entity:
            entities.add(entity)
    wb.close()
    return rows, entities


def count_internet_doctors(
    detail_rows: list[dict[str, str]],
    internet_rows: list[dict[str, Any]],
) -> list[int]:
    """按模板行顺序：执业医院名 = 实体医院 的精确匹配计数（去重医生 cert）。"""
    by_entity: dict[str, set[str]] = defaultdict(set)
    for row in detail_rows:
        hospital = str(row.get("执业医院") or "").strip()
        cert = normalize_cert(row.get("执业证书编码"))
        name = row.get("姓名", "")
        doc_key = cert or name
        if hospital:
            by_entity[hospital].add(doc_key)
    counts: list[int] = []
    for inet in internet_rows:
        entity = str(inet.get("实体医院") or "").strip()
        counts.append(len(by_entity.get(entity, set())) if entity else 0)
    return counts


def _clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return
    ws.delete_rows(start_row, ws.max_row - start_row + 1)


def _write_full_detail(ws, rows: list[dict[str, str]]) -> None:
    _clear_sheet_data(ws, 2)
    for col, h in enumerate(FULL_DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
    for r, row in enumerate(rows, 2):
        for col, h in enumerate(FULL_DETAIL_HEADERS, 1):
            c = ws.cell(row=r, column=col, value=row.get(h, ""))
            c.alignment = _CENTER


def _write_doctor_stats_sheet(ws, doctor_stats: dict[str, DoctorStat]) -> None:
    _clear_sheet_data(ws, 2)
    headers = ("医生档案", "医生姓名", "医生团队", "档案状态", "执业医院数", "互联网执业医院数")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.alignment = _CENTER
    ordered = sorted(doctor_stats.values(), key=lambda d: (d.name, d.cert))
    for r, doc in enumerate(ordered, 2):
        ws.cell(row=r, column=1, value=doc.file_id).alignment = _CENTER
        ws.cell(row=r, column=2, value=doc.name).alignment = _CENTER
        ws.cell(row=r, column=3, value=doc.team).alignment = _CENTER
        ws.cell(row=r, column=4, value=doc.status).alignment = _CENTER
        ws.cell(row=r, column=5, value=doc.hospital_count).alignment = _CENTER
        ws.cell(row=r, column=6, value=doc.internet_count).alignment = _CENTER

    coop = sum(1 for d in doctor_stats.values() if d.status == "启用")
    non_coop = sum(1 for d in doctor_stats.values() if d.status == "停用")
    total = len(doctor_stats)

    # 右侧汇总区 I4:M22
    ws["I4"] = "医生档案状态"
    ws["I5"] = "合作医生数"
    ws["J5"] = "不合作医生数"
    ws["I6"] = coop
    ws["J6"] = non_coop
    ws["I8"] = "合作-医生执业医院数"
    ws["L8"] = "合作-互联网医院执业数"
    ws["I9"] = "医院数"
    ws["J9"] = "医生数"
    ws["K9"] = "占比"
    ws["L9"] = "医生数"
    ws["M9"] = "占比"

    bin_h: Counter[str] = Counter()
    bin_i: Counter[str] = Counter()
    for doc in doctor_stats.values():
        bin_h[hospital_count_bin(doc.hospital_count)] += 1
        bin_i[hospital_count_bin(doc.internet_count)] += 1

    for i, label in enumerate(HOSPITAL_COUNT_BIN_ORDER):
        row = 10 + i
        ws.cell(row=row, column=9, value=label)
        h_cnt = bin_h.get(label, 0)
        i_cnt = bin_i.get(label, 0)
        ws.cell(row=row, column=10, value=h_cnt)
        ws.cell(row=row, column=11, value=h_cnt / total if total else 0)
        ws.cell(row=row, column=12, value=i_cnt)
        ws.cell(row=row, column=13, value=i_cnt / total if total else 0)
        for col in range(9, 14):
            ws.cell(row=row, column=col).alignment = _CENTER
            if col in (11, 13):
                ws.cell(row=row, column=col).number_format = "0.00%"

    ws["I22"] = "汇总"
    ws["J22"] = total
    ws["K22"] = 1
    ws["L22"] = total
    ws["M22"] = 1
    for coord in ("I4", "I5", "J5", "I6", "J6", "I8", "L8", "I9", "J9", "K9", "L9", "M9", "I22", "J22", "K22", "L22", "M22"):
        ws[coord].alignment = _CENTER
    ws["K22"].number_format = "0.00%"
    ws["M22"].number_format = "0.00%"


def _write_hospital_overlap(ws, overlap: list[tuple[str, int, int, int]]) -> None:
    _clear_sheet_data(ws, 2)
    headers = ("执业医院", "启用", "停用", "总计")
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).alignment = _CENTER
    for r, (name, en, dis, total) in enumerate(overlap, 2):
        ws.cell(row=r, column=1, value=name).alignment = _CENTER
        ws.cell(row=r, column=2, value=en).alignment = _CENTER
        ws.cell(row=r, column=3, value=dis).alignment = _CENTER
        ws.cell(row=r, column=4, value=total).alignment = _CENTER


def _update_internet_counts(ws, counts: list[int]) -> None:
    # 第 8 列「执业医生数」
    for i, cnt in enumerate(counts):
        cell = ws.cell(row=2 + i, column=8, value=cnt)
        cell.alignment = _CENTER


def build_report_workbook(
    *,
    detail_rows: list[dict[str, str]],
    template_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    internet_rows, internet_entities = read_internet_entities(template_path)
    doctor_stats = aggregate_doctors(detail_rows, internet_entities)
    overlap = aggregate_hospital_overlap(detail_rows)
    internet_counts = count_internet_doctors(detail_rows, internet_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    _write_full_detail(wb["全量明细"], detail_rows)
    _write_doctor_stats_sheet(wb["医生执业医院数"], doctor_stats)
    _write_hospital_overlap(wb["医院重叠数"], overlap)
    _update_internet_counts(wb["互联网医院重叠数"], internet_counts)

    wb.save(output_path)

    matched = sum(1 for r in detail_rows if r.get("档案编号"))
    return {
        "detailRows": len(detail_rows),
        "doctors": len(doctor_stats),
        "lianouMatchedRows": matched,
        "lianouMissRows": len(detail_rows) - matched,
        "hospitals": len(overlap),
        "coopDoctors": sum(1 for d in doctor_stats.values() if d.status == "启用"),
        "nonCoopDoctors": sum(1 for d in doctor_stats.values() if d.status == "停用"),
        "output": str(output_path),
    }


def run_build_practice_hospital_report(
    *,
    config_path: Path,
    template_path: Path,
    output_path: Path,
    institution_xlsx: Path,
    lianou_records: list[dict[str, Any]],
) -> dict[str, Any]:
    institution_rows = read_institution_rows(institution_xlsx)
    detail_rows = enrich_detail_rows(institution_rows, lianou_records)
    return build_report_workbook(
        detail_rows=detail_rows,
        template_path=template_path,
        output_path=output_path,
    )
