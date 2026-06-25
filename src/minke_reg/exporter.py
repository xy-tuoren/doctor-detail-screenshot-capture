from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Same internal storage as UI .xls (Excel serial); apply date format on write.
_DATE_HEADERS = frozenset({"审核日期", "开始日期", "结束日期"})
_DATE_NUMBER_FORMAT = "yyyy-mm-dd hh:mm:ss"


def _excel_serial_to_datetime(value: Any) -> datetime | None:
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    if serial < 1 or serial > 100_000:
        return None
    base = datetime(1899, 12, 30)
    return base + timedelta(days=serial)


def _cell_for_export(header: str, value: Any) -> Any:
    if header in _DATE_HEADERS and value not in ("", None):
        dt = _excel_serial_to_datetime(value)
        if dt is not None:
            return dt
    return value


from src.institution_export.paths import EXPORT_REG_API_DIR


def default_output_path(cfg: dict, label: str) -> Path:
    root = Path(__file__).resolve().parent.parent.parent
    out_dir = root / str(cfg.get("outputDir", f"exports/{EXPORT_REG_API_DIR}"))
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return out_dir / f"{label}导出-{stamp}.xlsx"


def save_reg_xlsx(records: list[dict[str, str]], output_path: Path) -> None:
    save_reg_workbook({"Sheet1": records}, output_path)


def save_reg_workbook(sheets: dict[str, list[dict[str, str]]], output_path: Path) -> int:
    wb = Workbook()
    wb.remove(wb.active)
    total = 0
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        date_col_indexes = [i for i, h in enumerate(headers, start=1) if h in _DATE_HEADERS]
        for row in rows:
            ws.append([_cell_for_export(h, row.get(h, "")) for h in headers])
            total += 1
        for idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 18
        for col_idx in date_col_indexes:
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row_idx}"]
                if isinstance(cell.value, datetime):
                    cell.number_format = _DATE_NUMBER_FORMAT
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return total
