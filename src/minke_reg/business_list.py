"""业务办理列表（独立流程，不属于核对/采图/写回主线）。

通过机构端 SOAP ``SearchListOfBusiness`` 拉取「业务办理」相关列表，
对应客户端菜单：

- aAudit=0：待机构确认（等待机构确认）
- aAudit=1：正在申办（机构已确认 / 等待行政部门审批）
- aAudit=2：已完成审批

可选字典：``GetBusinessType``（医师申请业务办理的业务类型清单）。
"""
from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.minke_reg.constants import DEFAULT_MINKE_REG_CONFIG, NS_DOCTOR_UNIT
from src.minke_reg.session import login_minke_reg
from src.minke_reg.soap import SoapClient, extract_error_message

# UI「业务办理」三档；sheet 名兼顾可读性与 aAudit 对照
AUDIT_SPECS: tuple[tuple[int, str, str], ...] = (
    (0, "待机构确认", "等待机构确认"),
    (1, "正在申办", "机构已确认 / 待行政部门审批"),
    (2, "已完成审批", "已完成（含行政部门审批时间）"),
)

# 中文导出列（其余原始字段追加在后）
EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Doctor_Name_Add", "姓名"),
    ("Doctor_IDCard_Add", "身份证号"),
    ("BisType_Name", "业务类型"),
    ("Create_Time", "申请日期"),
    ("FlagUnitAuditDesc", "机构审核状态"),
    ("FlagOrganAuditDesc", "行政部门审核状态"),
    ("ApprovalTime_Unit", "机构审核时间"),
    ("ApprovalTime_Org", "行政部门审核时间"),
    ("Approval_UnitName", "确认机构"),
    ("Approval_OrgName", "审批机关"),
    ("Bis_Code", "业务编号"),
    ("Bis_ID", "Bis_ID"),
    ("Bis_GID", "Bis_GID"),
    ("Doctor_GID", "Doctor_GID"),
    ("IsSubmit", "已提交"),
    ("LoginFromName", "来源"),
    ("RegCanel_Memo", "注销原因"),
)


def _load_reg_cfg(config_path: Path | str | None = None) -> dict[str, Any]:
    path = Path(config_path) if config_path else Path("config.json")
    cfg = json.loads(path.read_text(encoding="utf-8-sig"))
    reg = {**DEFAULT_MINKE_REG_CONFIG, **cfg.get("minkeRegApi", {})}
    reg["loginUser"] = cfg.get("loginUser")
    reg["loginPassword"] = cfg.get("loginPassword")
    return reg


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_dataset_any_rows(soap_xml: str, *, min_children: int = 2) -> list[dict[str, str]]:
    """从 SOAP DataSet 响应中解析最常见的数据行（跳过 field/selector 噪声）。"""
    import xml.etree.ElementTree as ET

    root = ET.fromstring(soap_xml)
    cands: list[tuple[str, dict[str, str]]] = []
    cnt: Counter[str] = Counter()
    for elem in root.iter():
        children = list(elem)
        if len(children) < min_children:
            continue
        leafish = sum(1 for c in children if len(list(c)) == 0)
        if leafish < min_children:
            continue
        row = {
            _local(c.tag): (c.text or "").strip()
            for c in children
            if len(list(c)) == 0
        }
        if len(row) < min_children:
            continue
        if set(row.keys()) <= {"field", "selector"}:
            continue
        name = _local(elem.tag)
        cnt[name] += 1
        cands.append((name, row))
    if not cnt:
        return []
    best = cnt.most_common(1)[0][0]
    return [r for t, r in cands if t == best]


def _iso_short(value: str) -> str:
    if not value:
        return ""
    s = value.replace("T", " ")
    if "+" in s:
        s = s.split("+", 1)[0]
    if "." in s:
        head, _rest = s.split(".", 1)
        if " " in head or len(head) >= 10:
            s = head
    return s.strip()


def search_list_of_business(
    reg: dict[str, Any],
    session,
    *,
    audit: int,
    months: int = 1,
    name: str = "",
    busi_type: str = "",
    id_card: str = "",
) -> list[dict[str, str]]:
    """调用 SearchListOfBusiness，返回原始行（日期字段已缩短显示）。"""
    client = SoapClient(
        str(reg["docUnitServiceUrl"]),
        NS_DOCTOR_UNIT,
        timeout=int(reg.get("requestTimeoutSeconds", 120)),
    )
    resp = client.call_operation(
        "SearchListOfBusiness",
        {
            "aAudit": int(audit),
            "aBusiType": busi_type or "",
            "aIDCard": id_card or "",
            "aMonths": int(months),
            "aName": name or "",
        },
        header_xml=session.doctor_header(),
    )
    err = extract_error_message(resp)
    if err:
        raise RuntimeError(f"SearchListOfBusiness(aAudit={audit}): {err}")
    rows = parse_dataset_any_rows(resp)
    for row in rows:
        for key in ("Create_Time", "ApprovalTime_Unit", "ApprovalTime_Org"):
            if key in row:
                row[key] = _iso_short(row[key])
    return rows


def fetch_business_types(reg: dict[str, Any], session) -> list[dict[str, str]]:
    """调用 GetBusinessType，返回业务类型字典（医师申请业务办理用）。"""
    client = SoapClient(
        str(reg["docUnitServiceUrl"]),
        NS_DOCTOR_UNIT,
        timeout=int(reg.get("requestTimeoutSeconds", 120)),
    )
    resp = client.call_operation(
        "GetBusinessType",
        {"aMd5Str": ""},
        header_xml=session.doctor_header(),
    )
    err = extract_error_message(resp)
    if err:
        raise RuntimeError(f"GetBusinessType: {err}")
    return parse_dataset_any_rows(resp)


def default_business_list_output_path(
    artifacts_dir: Path | str | None = None,
    *,
    months: int = 1,
) -> Path:
    root = Path(artifacts_dir) if artifacts_dir else Path("workspace") / "artifacts"
    root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return root / f"业务办理列表_{stamp}_近{months}月.xlsx"


def _write_business_sheet(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        ws.append(["（无数据）"])
        return
    preferred = [k for k, _ in EXPORT_COLUMNS]
    alias = {k: cn for k, cn in EXPORT_COLUMNS}
    extra = sorted({k for r in rows for k in r.keys()} - set(preferred))
    headers = preferred + extra
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    ws.append([alias.get(h, h) for h in headers])
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for idx, h in enumerate(headers, 1):
        label = alias.get(h, h)
        ws.column_dimensions[get_column_letter(idx)].width = min(28, max(10, len(label) + 4))
    ws.freeze_panes = "A2"
    last = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last}{max(1, len(rows) + 1)}"


def _write_types_sheet(ws, rows: list[dict[str, str]]) -> None:
    headers = [
        ("BisType_ID", "类型ID"),
        ("BisType_Code", "类型编码"),
        ("BisType_Name", "业务类型"),
        ("Has_Detailed", "有详情"),
        ("Summary_Txt", "摘要"),
    ]
    bold = Font(bold=True)
    ws.append([cn for _, cn in headers])
    for cell in ws[1]:
        cell.font = bold
    keys = [k for k, _ in headers]
    for row in rows:
        ws.append([row.get(k, "") for k in keys])
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.column_dimensions["E"].width = 48
    ws.freeze_panes = "A2"


def fetch_business_lists_to_xlsx(
    *,
    config_path: Path | str | None = None,
    reg_cfg: dict[str, Any] | None = None,
    audits: list[int] | None = None,
    months: int = 1,
    name: str = "",
    busi_type: str = "",
    id_card: str = "",
    include_types: bool = False,
    output_path: Path | str | None = None,
    on_progress: Callable[[str, int, int], None] | None = None,
) -> tuple[dict[int, int], Path]:
    """拉取业务办理列表并写 xlsx。

    返回 ``({aAudit: 条数}, 输出路径)``。
    """
    reg = reg_cfg if reg_cfg is not None else _load_reg_cfg(config_path)
    session = login_minke_reg(reg)
    audit_list = list(audits) if audits is not None else [a for a, _, _ in AUDIT_SPECS]
    for a in audit_list:
        if a not in (0, 1, 2):
            raise ValueError(f"aAudit 仅支持 0/1/2，收到: {a}")

    label_by_audit = {a: title for a, title, _ in AUDIT_SPECS}
    counts: dict[int, int] = {}
    sheet_rows: dict[int, list[dict[str, str]]] = {}

    for audit in audit_list:
        rows = search_list_of_business(
            reg,
            session,
            audit=audit,
            months=months,
            name=name,
            busi_type=busi_type,
            id_card=id_card,
        )
        sheet_rows[audit] = rows
        counts[audit] = len(rows)
        if on_progress:
            on_progress(label_by_audit.get(audit, f"aAudit={audit}"), audit, len(rows))

    types_rows: list[dict[str, str]] = []
    if include_types:
        types_rows = fetch_business_types(reg, session)
        if on_progress:
            on_progress("业务类型字典", -1, len(types_rows))

    out = Path(output_path) if output_path else default_business_list_output_path(months=months)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # 汇总
    ws_sum = wb.create_sheet("汇总", 0)
    ws_sum.append(["sheet", "aAudit", "条数", "时间段(月)", "姓名筛选", "业务类型筛选", "拉取时间"])
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for audit in audit_list:
        ws_sum.append(
            [
                label_by_audit.get(audit, f"aAudit={audit}"),
                audit,
                counts.get(audit, 0),
                months,
                name or "",
                busi_type or "",
                stamp,
            ]
        )
    ws_sum.append([])
    ws_sum.append(["接口", "SearchListOfBusiness"])
    ws_sum.append(["说明", "aAudit=0 待机构确认；1 正在申办；2 已完成审批"])
    ws_sum.append(["对应 UI", "业务办理 → 医师正在申办的业务列表 / 已完成审批业务列表"])

    for audit in audit_list:
        title = label_by_audit.get(audit, f"aAudit{audit}")
        ws = wb.create_sheet(f"{title}_aAudit{audit}"[:31])
        _write_business_sheet(ws, sheet_rows.get(audit) or [])

    if include_types:
        ws_t = wb.create_sheet("业务类型字典")
        _write_types_sheet(ws_t, types_rows)

    wb.save(out)
    return counts, out
