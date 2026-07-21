"""医生执业医院明细获取（独立流程，不属于核对/采图/写回主线）。

通过机构端 SOAP 接口获取每个医生在所有省份的主执业 + 多执业备案信息，
输出为 Excel 明细表。数据来源：

- ``DoctorUnitGetListForOther`` (st=1/8/9/10/11)：获取注册行（主执业 + 外省注册）
- ``GetRegDetailForUnit``：获取注册行详情（审批日期、医院地址、省份）
- ``GetMutiRegListByRegisterId``：获取每个注册行下属的多执业备案（含起止日期）
- ``make_electronic_license_url``：本地生成电子证照预览 URL（AES-128-CBC）
- HTTP GET 电子证照 URL 解析 <title> 判断是否已申领

医生筛选：主执业为本院（st=1）∪ 多执业含本院（st=8），拿全这些医生在所有
省份的主执业 + 多执业备案信息，全局去重。
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import re
import threading
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.minke_reg.constants import (
    DEFAULT_MINKE_REG_CONFIG,
    MAIN_ROW_TAGS,
    MULTI_ROW_TAGS,
    NS_DOCTOR_UNIT,
)
from src.minke_reg.dataset import parse_dataset_rows
from src.minke_reg.doctor_detail import (
    _detail_client,
    format_audit_date,
)
from src.minke_reg.elec_license import check_elec_applied, make_electronic_license_url
from src.minke_reg.session import login_minke_reg
from src.minke_reg.soap import SoapClient, extract_error_message

HOME_PROVINCE = "广东省"
HOME_UNIT = "莲藕健康医院"

# 本院（莲藕健康医院）固定地址缓存：st=1 detail 的 UnitAddress。
# st=8（外院医生来本院多执业）行的"医院地址"应为本院地址，而非外院地址。
_HOME_ADDR_LOCK = threading.Lock()
_HOME_UNIT_ADDRESS: str = ""


def _resolve_home_address(detail_cache: dict[tuple[str, str], dict[str, str]]) -> str:
    """返回本院（莲藕健康医院）地址。优先用已缓存值；否则扫描 detail_cache
    找 UnitName=HOME_UNIT 的第一个 UnitAddress，原子缓存后返回。"""
    global _HOME_UNIT_ADDRESS
    with _HOME_ADDR_LOCK:
        if _HOME_UNIT_ADDRESS:
            return _HOME_UNIT_ADDRESS
        for d in detail_cache.values():
            if d.get("UnitName") == HOME_UNIT and d.get("UnitAddress"):
                _HOME_UNIT_ADDRESS = d["UnitAddress"]
                return _HOME_UNIT_ADDRESS
    return ""


# 中国 34 个省级行政区标准名。用于规整 detail.AreaName 可能返回的"省+市+区"拼串。
_PROVINCE_NAMES: tuple[str, ...] = (
    "北京市", "天津市", "河北省", "山西省", "内蒙古自治区", "辽宁省", "吉林省",
    "黑龙江省", "上海市", "江苏省", "浙江省", "安徽省", "福建省", "江西省",
    "山东省", "河南省", "湖北省", "湖南省", "广东省", "广西壮族自治区", "海南省",
    "重庆市", "四川省", "贵州省", "云南省", "西藏自治区", "陕西省", "甘肃省",
    "青海省", "宁夏回族自治区", "新疆维吾尔自治区", "台湾省", "香港特别行政区",
    "澳门特别行政区",
)
_PROVINCE_PREFIXES: tuple[tuple[str, str], ...] = tuple(
    (p[:2], p) for p in _PROVINCE_NAMES
)


def _normalize_province(raw: str) -> str:
    """规整省份字段：若 raw 已是标准省级名则原样返回；否则按前缀匹配标准名
    （如"广东揭阳市榕城区"→"广东省"）；匹配不到则原样返回。"""
    if not raw:
        return ""
    s = raw.strip()
    if s in _PROVINCE_NAMES:
        return s
    for prefix, name in _PROVINCE_PREFIXES:
        if s.startswith(prefix):
            return name
    return s

DETAIL_HEADERS: tuple[str, ...] = (
    "姓名",
    "身份证号",
    "执业证书编码",
    "性别",
    "医师类别",
    "医师级别",
    "执业范围",
    "任职资格",
    "审批日期",
    "开始日期",
    "结束日期",
    "是否主执业机构",
    "是否省外",
    "执业医院",
    "医院地址",
    "省份",
    "数据来源",
)

# (searchType, 数据来源标签, 内部 kind)
SEARCH_SPECS: tuple[tuple[int, str, str], ...] = (
    (1, "主执业机构在本院", "main_at_home"),
    (8, "外院在本院多执业", "multi_at_home"),
    (9, "本院多机构备案", "home_multi_filing"),
    (10, "主执业在本院-外省注册", "main_out_province"),
    (11, "多机构在本院-外省注册", "multi_out_province"),
)


def _load_reg_cfg(config_path: Path | str | None = None) -> dict:
    path = Path(config_path) if config_path else Path("config.json")
    cfg = json.loads(path.read_text(encoding="utf-8-sig"))
    reg = {**DEFAULT_MINKE_REG_CONFIG, **cfg.get("minkeRegApi", {})}
    reg["loginUser"] = cfg.get("loginUser")
    reg["loginPassword"] = cfg.get("loginPassword")
    return reg


def _iso_date(value: str) -> str:
    if not value:
        return ""
    try:
        return _dt.datetime.fromisoformat(value.replace("+08:00", "")).strftime("%Y-%m-%d")
    except ValueError:
        return value


def _excel_date(value: str | float) -> str:
    if value == "" or value is None:
        return ""
    if isinstance(value, (int, float)):
        base = _dt.datetime(1899, 12, 30)
        return (base + _dt.timedelta(days=float(value))).strftime("%Y-%m-%d")
    return _iso_date(str(value))


def _fetch_list_for_other(
    reg: dict, session, search_type: int, cache: dict[int, list[dict[str, str]]] | None = None,
) -> list[dict[str, str]]:
    if cache is not None and search_type in cache:
        return cache[search_type]
    client = SoapClient(
        str(reg["docUnitServiceUrl"]),
        NS_DOCTOR_UNIT,
        timeout=int(reg.get("requestTimeoutSeconds", 120)),
    )
    resp = client.call_operation(
        "DoctorUnitGetListForOther",
        {"aSeachType": search_type, "aMd5Str": ""},
        header_xml=session.doctor_header(),
    )
    err = extract_error_message(resp)
    if err:
        raise RuntimeError(f"DoctorUnitGetListForOther({search_type}): {err}")
    rows = parse_dataset_rows(resp, MULTI_ROW_TAGS)
    if not rows:
        rows = parse_dataset_rows(resp, MAIN_ROW_TAGS)
    if cache is not None:
        cache[search_type] = rows
    return rows


# 一次 GetRegDetailForUnit 调用解析全部需要的字段（含地址/医院名/省份/审批日期），
# 避免 _rows_for_doctor 里对同一 GID 重复调用。
_FULL_DETAIL_FIELDS: tuple[str, ...] = (
    "IDCard",
    "WorkLicenceCode",
    "CPETLicenceCode",
    "PostCpetName",
    "WorkCpetName",
    "UnitAddress",
    "UnitName",
    "AreaName",
    "LastPassDate",
)


def _first_tag(soap_xml: str, tag: str) -> str:
    m = re.search(rf"<{re.escape(tag)}>([^<]*)</", soap_xml, re.I)
    return m.group(1).strip() if m else ""


def _detail_map_raw(reg: dict, session, doctor_gid: str, register_gid: str) -> dict[str, str]:
    """一次 GetRegDetailForUnit 调用，按 RegisterID 精确匹配出该注册行的字段。

    GetRegDetailForUnit 返回该医生**所有注册行**的合集（多个 UnitName/UnitAddress/
    RegisterID 节点），必须按传入的 register_gid 过滤出对应那一行，否则会取到
    列表第一行的医院名配到另一行的地址，造成错配。
    """
    if not doctor_gid or not register_gid:
        raise ValueError("empty doctor_gid/register_gid")
    raw = _detail_client(reg).call_operation(
        "GetRegDetailForUnit",
        {
            "aDoctorId": doctor_gid,
            "aRegisterID": register_gid,
            "aIdCard": "",
            "aPhoto": "false",
        },
        header_xml=session.doctor_header(),
    )
    err = extract_error_message(raw)
    if err:
        raise RuntimeError(f"GetRegDetailForUnit error: {err}")
    root = ET.fromstring(raw)
    parent = {c: p for p in root.iter() for c in p}
    # RegisterID 会出现在多个容器：tRegAuditingMind(仅含 RegisterID)、
    # Register(注册主表，含 UnitName/UnitAddress/LastPassDate/PostCpetName 全字段)、
    # PreRegister(预注册，同全字段)。必须取 Register/PreRegister，按 RegisterID 匹配，
    # 否则多个 tExperience 的 UnitName 会被 _first_tag 误取为列表第一个。
    candidates: list[tuple[str, object]] = []
    for elem in root.iter():
        if (
            elem.tag.rsplit("}", 1)[-1] == "RegisterID"
            and (elem.text or "").strip() == register_gid
            and elem in parent
        ):
            cont = parent[elem]
            cont_tag = cont.tag.rsplit("}", 1)[-1]
            if cont_tag in ("Register", "PreRegister"):
                candidates.append((cont_tag, cont))
    for want in ("Register", "PreRegister"):
        for cont_tag, cont in candidates:
            if cont_tag == want:
                fields: dict[str, str] = {}
                for c in cont:
                    tag = c.tag.rsplit("}", 1)[-1]
                    if tag in _FULL_DETAIL_FIELDS and c.text:
                        fields.setdefault(tag, c.text.strip())
                if fields:
                    return fields
    # 回退：未找到匹配容器，用首个匹配（极少触发）
    return {f: _first_tag(raw, f) for f in _FULL_DETAIL_FIELDS}


def _detail_map(reg: dict, session, doctor_gid: str, register_gid: str) -> dict[str, str]:
    """兼容包装：失败返回 {}（串行模式用）。"""
    try:
        return _detail_map_raw(reg, session, doctor_gid, register_gid)
    except Exception:
        return {}


_MUTI_REG_CACHE: dict[str, list[dict[str, str]]] = {}
_MUTI_REG_LOCK = threading.Lock()


def _fetch_muti_reg_list(reg: dict, session, register_gid: str) -> list[dict[str, str]]:
    """调用 GetMutiRegListByRegisterId 获取某注册行下属的多执业备案列表。失败返回 [] 但不缓存。"""
    if not register_gid:
        return []
    with _MUTI_REG_LOCK:
        if register_gid in _MUTI_REG_CACHE:
            return _MUTI_REG_CACHE[register_gid]
    try:
        raw = _detail_client(reg).call_operation(
            "GetMutiRegListByRegisterId",
            {"aRegisterId": register_gid},
            header_xml=session.doctor_header(),
        )
        rows: list[dict[str, str]] = []
        for elem in ET.fromstring(raw).iter():
            if elem.tag.rsplit("}", 1)[-1] == "vRegMutiList":
                row = {c.tag.rsplit("}", 1)[-1]: (c.text or "").strip() for c in elem}
                if row:
                    rows.append(row)
    except Exception:
        return []  # 不缓存，预取重试/补取可再试
    with _MUTI_REG_LOCK:
        _MUTI_REG_CACHE[register_gid] = rows
    return rows


def _base_from_list(row: dict[str, str], detail: dict[str, str]) -> dict[str, str]:
    post = detail.get("PostCpetName") or detail.get("WorkCpetName") or ""
    return {
        "姓名": row.get("Doctor_Name", ""),
        "身份证号": row.get("IDCard") or detail.get("IDCard", ""),
        "执业证书编码": detail.get("WorkLicenceCode", "") or row.get("WorkLicenceCode", ""),
        "性别": row.get("Sex", ""),
        "医师类别": row.get("Doctor_SortName", ""),
        "医师级别": row.get("Doctor_LevelName", ""),
        "执业范围": row.get("Subject_Name", ""),
        "任职资格": post,
    }


def _profile_from_main(
    reg: dict, session, name: str, list_cache: dict[int, list[dict[str, str]]] | None = None,
) -> dict[str, str]:
    rows = _fetch_list_for_other(reg, session, 1, list_cache)
    matched = [r for r in rows if r.get("Doctor_Name") == name]
    if not matched:
        return {}
    row = matched[0]
    detail = _detail_map(
        reg, session, str(row.get("Doctor_GID", "")), str(row.get("Doctor_RegisterGID", ""))
    )
    return _base_from_list(row, detail)


def _build_name_index(
    list_cache: dict[int, list[dict[str, str]]],
) -> dict[str, dict[int, list[dict[str, str]]]]:
    """按姓名建索引：name -> {search_type: [matched rows]}，避免 3459×5 次全表扫描。"""
    index: dict[str, dict[int, list[dict[str, str]]]] = {}
    for st, rows in list_cache.items():
        for r in rows:
            name = (r.get("Doctor_Name") or "").strip()
            if not name:
                continue
            index.setdefault(name, {}).setdefault(st, []).append(r)
    return index


_DETAIL_CACHE: dict[tuple[str, str], dict[str, str]] = {}
_DETAIL_CACHE_LOCK = threading.Lock()


def _detail_cached(reg, session, doctor_gid: str, register_gid: str) -> dict[str, str]:
    """线程安全的 detail 缓存读取/填充。失败不缓存，便于后续重试。"""
    key = (doctor_gid, register_gid)
    with _DETAIL_CACHE_LOCK:
        if key in _DETAIL_CACHE:
            return _DETAIL_CACHE[key]
    try:
        val = _detail_map_raw(reg, session, doctor_gid, register_gid)
    except Exception:
        return {}  # 不缓存
    with _DETAIL_CACHE_LOCK:
        _DETAIL_CACHE[key] = val
    return val


def _collect_prefetch_keys(
    name_index: dict[str, dict[int, list[dict[str, str]]]],
    names: list[str],
) -> tuple[list[tuple[str, str]], list[str]]:
    """收集需要预取的 detail key 和 muti register_id。"""
    detail_keys: set[tuple[str, str]] = set()
    muti_reg_ids: set[str] = set()
    for name in names:
        per_st = name_index.get(name) or {}
        for st, _, kind in SEARCH_SPECS:
            for r in per_st.get(st, []):
                dg = str(r.get("Doctor_GID", ""))
                rg = str(r.get("Doctor_RegisterGID", ""))
                if dg and rg:
                    detail_keys.add((dg, rg))
                if kind in ("main_out_province", "multi_out_province") and rg:
                    muti_reg_ids.add(rg)
    return sorted(detail_keys), sorted(muti_reg_ids)


def _prefetch_details(
    reg, session, keys: list[tuple[str, str]], workers: int, on_progress=None,
) -> dict | None:
    """并行预取 detail，task 内重试 2 次；会话失效自动重登录；失败 key 用 4 线程低并发补一轮。

    返回 ``{"session": ...}`` holder（session 可能已被重登录更新），无 key 时返回 None。
    """
    if not keys:
        return None
    import time

    holder = {"session": session}
    relogin_lock = threading.Lock()
    last_relogin_ts = [0.0]
    RELOGIN_MIN_INTERVAL = 10.0  # 秒；避免多线程同时失效时重复重登录

    def maybe_relogin() -> None:
        """线程安全重登录：10 秒内只重登录一次，其他线程复用新 session。"""
        with relogin_lock:
            now = time.time()
            if now - last_relogin_ts[0] < RELOGIN_MIN_INTERVAL:
                return
            last_relogin_ts[0] = now
            holder["session"] = login_minke_reg(reg)

    def task(dg, rg):
        key = (dg, rg)
        for attempt in range(3):  # 首次 + 2 重试
            try:
                val = _detail_map_raw(reg, holder["session"], dg, rg)
                with _DETAIL_CACHE_LOCK:
                    _DETAIL_CACHE[key] = val
                return True
            except Exception as exc:
                msg = str(exc)
                if "非法的用户身份" in msg:
                    maybe_relogin()
                    continue  # 用新 session 重试，不计入 attempt
                if attempt < 2:
                    time.sleep(0.3 * (attempt + 1))
        return False

    done = 0
    total = len(keys)
    failed: list[tuple[str, str]] = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(task, dg, rg): (dg, rg) for dg, rg in keys}
        for fut in as_completed(futs):
            ok = False
            try:
                ok = fut.result()
            except Exception:
                pass
            if not ok:
                failed.append(futs[fut])
            done += 1
            if on_progress and (done == 1 or done == total or done % 50 == 0):
                on_progress("detail", done, total)

    # 失败的 key 低并发补取（4 线程，服务端压力小，成功率高）
    if failed:
        if on_progress:
            on_progress("detail-retry", 0, len(failed))
        retry_done = 0
        retry_lock = threading.Lock()

        def retry_progress(_ok):
            nonlocal retry_done
            with retry_lock:
                retry_done += 1
                if on_progress and (retry_done == 1 or retry_done == len(failed) or retry_done % 100 == 0):
                    on_progress("detail-retry", retry_done, len(failed))

        with ThreadPoolExecutor(max_workers=min(4, len(failed))) as pool:
            futs = {pool.submit(task, dg, rg): dg for dg, rg in failed}
            for fut in as_completed(futs):
                try:
                    fut.result()
                except Exception:
                    pass
                retry_progress(True)
    return holder


def _prefetch_muti(
    reg, session, reg_ids: list[str], workers: int, on_progress=None,
) -> dict | None:
    """并行预取 muti，会话失效自动重登录；失败 id 用 4 线程低并发补一轮。

    返回 ``{"session": ...}`` holder，无 id 时返回 None。
    """
    if not reg_ids:
        return None
    import time

    holder = {"session": session}
    relogin_lock = threading.Lock()
    last_relogin_ts = [0.0]
    RELOGIN_MIN_INTERVAL = 10.0

    def maybe_relogin() -> None:
        with relogin_lock:
            now = time.time()
            if now - last_relogin_ts[0] < RELOGIN_MIN_INTERVAL:
                return
            last_relogin_ts[0] = now
            holder["session"] = login_minke_reg(reg)

    def task(rid):
        for attempt in range(3):
            try:
                raw = _detail_client(reg).call_operation(
                    "GetMutiRegListByRegisterId",
                    {"aRegisterId": rid},
                    header_xml=holder["session"].doctor_header(),
                )
                rows: list[dict[str, str]] = []
                for elem in ET.fromstring(raw).iter():
                    if elem.tag.rsplit("}", 1)[-1] == "vRegMutiList":
                        row = {c.tag.rsplit("}", 1)[-1]: (c.text or "").strip() for c in elem}
                        if row:
                            rows.append(row)
                with _MUTI_REG_LOCK:
                    _MUTI_REG_CACHE[rid] = rows
                return True
            except Exception as exc:
                msg = str(exc)
                if "非法的用户身份" in msg:
                    maybe_relogin()
                    continue
                if attempt < 2:
                    time.sleep(0.3 * (attempt + 1))
        return False

    done = 0
    total = len(reg_ids)
    failed: list[str] = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(task, rid): rid for rid in reg_ids}
        for fut in as_completed(futs):
            ok = False
            try:
                ok = fut.result()
            except Exception:
                pass
            if not ok:
                failed.append(futs[fut])
            done += 1
            if on_progress and (done == 1 or done == total or done % 50 == 0):
                on_progress("muti", done, total)

    if failed:
        if on_progress:
            on_progress("muti-retry", 0, len(failed))
        retry_done = 0
        retry_lock = threading.Lock()

        def retry_progress(_ok):
            nonlocal retry_done
            with retry_lock:
                retry_done += 1
                if on_progress and (retry_done == 1 or retry_done == len(failed) or retry_done % 100 == 0):
                    on_progress("muti-retry", retry_done, len(failed))

        with ThreadPoolExecutor(max_workers=min(4, len(failed))) as pool:
            futs = {pool.submit(task, rid): rid for rid in failed}
            for fut in as_completed(futs):
                try:
                    fut.result()
                except Exception:
                    pass
                retry_progress(True)
    return holder


def _assemble_rows_for_doctor(
    reg: dict,
    session,
    name: str,
    name_index: dict[str, dict[int, list[dict[str, str]]]],
    detail_cache: dict[tuple[str, str], dict[str, str]],
    muti_cache: dict[str, list[dict[str, str]]],
) -> list[dict[str, str]]:
    """纯内存组装：从缓存读 detail/muti，不再调 SOAP。"""
    out: list[dict[str, str]] = []
    per_st = name_index.get(name) or {}
    # profile 取 st=1 第一行的 detail（与原 _profile_from_main 口径一致）
    profile: dict[str, str] = {}
    home_addr = ""
    for r in (per_st.get(1) or []):
        dg, rg = str(r.get("Doctor_GID", "")), str(r.get("Doctor_RegisterGID", ""))
        d = detail_cache.get((dg, rg)) or {}
        profile = _base_from_list(r, d)
        if d.get("UnitAddress") and not home_addr:
            home_addr = d["UnitAddress"]
        break
    if not home_addr:
        home_addr = _resolve_home_address(detail_cache)

    for search_type, source_label, kind in SEARCH_SPECS:
        matched = per_st.get(search_type) or []
        if not matched:
            continue
        for row in matched:
            dg, rg = str(row.get("Doctor_GID", "")), str(row.get("Doctor_RegisterGID", ""))
            detail = detail_cache.get((dg, rg)) or {}
            base = _base_from_list(row, detail)
            for key in ("身份证号", "性别", "任职资格"):
                if not base.get(key) and profile.get(key):
                    base[key] = profile[key]

            if kind == "main_at_home":
                # 本院主执业必在 HOME_PROVINCE（广东省）。固定省份，避免 detail.AreaName
                # 返回"广东揭阳市榕城区"这类省+市+区拼串污染省份列。
                province = HOME_PROVINCE
                hospital = detail.get("UnitName") or row.get("Unit_Name", HOME_UNIT)
                addr = detail.get("UnitAddress", "")
                audit = (
                    row.get("LastApprovalTime")
                    or row.get("Add_UpdateApproval_Time")
                    or detail.get("LastPassDate", "")
                )
                out.append(
                    {
                        **base,
                        "审批日期": _excel_date(format_audit_date(audit))
                        if audit and "T" not in str(audit)
                        else _iso_date(str(audit)),
                        "开始日期": "",
                        "结束日期": "",
                        "是否主执业机构": "是",
                        "是否省外": "否",
                        "执业医院": hospital,
                        "医院地址": addr,
                        "省份": province,
                        "数据来源": source_label,
                    }
                )
            elif kind == "multi_at_home":
                # st=8：外院医生来本院多执业。执业医院=本院(MutiUnitName)，
                # 医院地址=本院地址（非外院 detail 地址）；备案审批日期接口不返回，留空。
                out.append(
                    {
                        **base,
                        "审批日期": "",
                        "开始日期": _iso_date(row.get("BeginDate", "")),
                        "结束日期": _iso_date(row.get("EndDate", "")),
                        "是否主执业机构": "否",
                        "是否省外": "否",
                        "执业医院": row.get("MutiUnitName", HOME_UNIT),
                        "医院地址": home_addr,
                        "省份": HOME_PROVINCE,
                        "数据来源": f"{source_label}（主执业:{row.get('Unit_Name','')}）",
                    }
                )
            elif kind == "home_multi_filing":
                # st=9：本院医生去外院多机构备案。执业医院=外院(MutiUnitName)，
                # 外院地址接口不返回（MutiUnitName 无 GID），留空；
                # 审批日期应为外院备案审批，接口不返回，留空（detail 是本院审批，与本案无关）。
                out.append(
                    {
                        **base,
                        "审批日期": "",
                        "开始日期": _iso_date(row.get("BeginDate", "")),
                        "结束日期": _iso_date(row.get("EndDate", "")),
                        "是否主执业机构": "否",
                        "是否省外": "否",
                        "执业医院": row.get("MutiUnitName", ""),
                        "医院地址": "",
                        "省份": HOME_PROVINCE,
                        "数据来源": source_label,
                    }
                )
            elif kind in ("main_out_province", "multi_out_province"):
                province = _normalize_province(detail.get("AreaName") or row.get("AreaName", ""))
                hospital = detail.get("UnitName") or row.get("Unit_Name", "")
                addr = detail.get("UnitAddress", "")
                audit_date = _iso_date(detail.get("LastPassDate", ""))
                out.append(
                    {
                        **base,
                        "审批日期": audit_date,
                        "开始日期": "",
                        "结束日期": "",
                        "是否主执业机构": "是" if kind == "main_out_province" else "否",
                        "是否省外": "是",
                        "执业医院": hospital,
                        "医院地址": addr,
                        "省份": province,
                        "数据来源": source_label,
                    }
                )
                for mr in (muti_cache.get(rg) or []):
                    if not mr.get("MutiUnitName"):
                        continue
                    m_province = _normalize_province(mr.get("AreaReg", "")) or province
                    out.append(
                        {
                            **base,
                            "审批日期": _iso_date(mr.get("MutiOrganDate", "")),
                            "开始日期": _iso_date(mr.get("MutiBeginDate", "")),
                            "结束日期": _iso_date(mr.get("MutiEndDate", "")),
                            "是否主执业机构": "否",
                            "是否省外": "是",
                            "执业医院": mr.get("MutiUnitName", ""),
                            "医院地址": "",
                            "省份": m_province,
                            "数据来源": f"多执业机构备案（主执业:{hospital}）",
                        }
                    )

    # 全局去重
    seen_muti: set[tuple] = set()
    deduped: list[dict[str, str]] = []
    for row in out:
        if row.get("是否主执业机构") == "否":
            key = (
                row.get("姓名", ""),
                row.get("执业医院", ""),
                row.get("开始日期", ""),
                row.get("结束日期", ""),
            )
            if key in seen_muti:
                continue
            seen_muti.add(key)
        deduped.append(row)
    return deduped


def _rows_for_doctor(
    reg: dict,
    session,
    name: str,
    list_cache: dict[int, list[dict[str, str]]] | None = None,
) -> list[dict[str, str]]:
    """获取单个医生在所有省份的主执业 + 多执业备案行（串行，兼容旧调用）。"""
    if list_cache is None:
        list_cache = {}
    name_index = _build_name_index(list_cache)
    detail_keys, muti_reg_ids = _collect_prefetch_keys(name_index, [name])
    for dg, rg in detail_keys:
        _detail_cached(reg, session, dg, rg)
    for rid in muti_reg_ids:
        _fetch_muti_reg_list(reg, session, rid)
    # 确保本院地址已解析：若当前医生非本院主执业(st=1)，主动预取一个本院行 detail
    if not _HOME_UNIT_ADDRESS:
        for r in (list_cache.get(1) or []):
            _detail_cached(reg, session, str(r.get("Doctor_GID", "")), str(r.get("Doctor_RegisterGID", "")))
            break
        _resolve_home_address(_DETAIL_CACHE)
    return _assemble_rows_for_doctor(
        reg, session, name, name_index, _DETAIL_CACHE, _MUTI_REG_CACHE
    )


def _default_practice_table_path(doctor_names: list[str]) -> Path:
    """默认输出路径：单医生带姓名后缀，多人统一为 医生执业医院.xlsx。"""
    base = Path("workspace/artifacts")
    if len(doctor_names) == 1:
        return base / f"医生执业医院_{doctor_names[0]}.xlsx"
    return base / "医生执业医院.xlsx"


def _write_xlsx(rows: list[dict[str, str]], path: Path, sheet_name: str = "明细") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    bold = Font(bold=True)
    for col_idx, h in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([row.get(h, "") for h in DETAIL_HEADERS])
    widths = {
        "姓名": 8, "身份证号": 20, "执业证书编码": 22, "性别": 6,
        "医师类别": 8, "医师级别": 10, "执业范围": 12, "任职资格": 10,
        "审批日期": 12,
        "开始日期": 12, "结束日期": 12, "是否主执业机构": 12,
        "是否省外": 8, "执业医院": 24, "医院地址": 36, "省份": 10,
        "数据来源": 22,
    }
    for col_idx, h in enumerate(DETAIL_HEADERS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(h, 12)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(DETAIL_HEADERS))
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def top_doctor_names(
    reg: dict, session, limit: int, list_cache: dict[int, list[dict[str, str]]] | None = None,
) -> list[str]:
    """从主执业在本院(st=1) + 多执业含本院(st=8) 交替取前 limit 名医生(去重)。"""
    rows1 = _fetch_list_for_other(reg, session, 1, list_cache)
    rows8 = _fetch_list_for_other(reg, session, 8, list_cache)
    names: list[str] = []
    seen: set[str] = set()
    i = 0
    while len(names) < limit and (i < len(rows1) or i < len(rows8)):
        for rows in (rows1, rows8):
            if i < len(rows):
                n = rows[i].get("Doctor_Name", "").strip()
                if n and n not in seen:
                    seen.add(n)
                    names.append(n)
                    if len(names) >= limit:
                        break
        i += 1
    return names


def all_doctor_names(
    list_cache: dict[int, list[dict[str, str]]],
) -> list[str]:
    """全量名单：主执业在本院(st=1) ∪ 多执业含本院(st=8) 全部去重姓名（保持稳定顺序）。"""
    seen: set[str] = set()
    names: list[str] = []
    for rows in (list_cache.get(1) or [], list_cache.get(8) or []):
        for r in rows:
            n = (r.get("Doctor_Name") or "").strip()
            if n and n not in seen:
                seen.add(n)
                names.append(n)
    return names


def default_workers(reg: dict | None = None) -> int:
    """I/O 密集型并发：按 CPU 核数×2 推算，封顶 16（避免单 SOAP session 限流）；config 可覆盖。"""
    if reg and reg.get("practiceTableWorkers"):
        return max(1, int(reg["practiceTableWorkers"]))
    cpus = os.cpu_count() or 4
    return max(4, min(cpus * 2, 16))


def fetch_practice_table(
    doctor_names: list[str],
    reg_cfg: dict | None = None,
    config_path: Path | str | None = None,
    output_path: Path | str | None = None,
    sheet_name: str = "明细",
    on_progress=None,
    list_cache: dict[int, list[dict[str, str]]] | None = None,
) -> tuple[int, Path, dict[str, int]]:
    """获取多名医生的执业医院明细并写入 Excel。

    返回 (总行数, 输出路径, 每医生行数)。
    可传入 list_cache 复用已拉取的全量列表（如 top_doctor_names 已填充 st=1/8）。
    """
    reg = reg_cfg or _load_reg_cfg(config_path)
    session = login_minke_reg(reg)

    # 跨医生共享全量列表缓存：批量开始时每个 searchType 只拉一次，
    # 后续医生从内存 filter，把列表调用从 6×N 降到 6 次。
    if list_cache is None:
        list_cache = {}
    if on_progress:
        on_progress(0, len(doctor_names), "预拉全量列表")
    for st, _label, _kind in SEARCH_SPECS:
        if st not in list_cache:
            _fetch_list_for_other(reg, session, st, list_cache)

    all_rows: list[dict[str, str]] = []
    per_doctor: dict[str, int] = {}
    for i, name in enumerate(doctor_names, 1):
        if on_progress:
            on_progress(i, len(doctor_names), name)
        rows: list[dict[str, str]] = []
        for attempt in range(2):
            try:
                rows = _rows_for_doctor(reg, session, name, list_cache)
                break
            except Exception as exc:
                msg = str(exc)
                if "非法的用户身份" in msg and attempt == 0:
                    session = login_minke_reg(reg)
                    list_cache.clear()
                    for st, _l, _k in SEARCH_SPECS:
                        _fetch_list_for_other(reg, session, st, list_cache)
                    continue
                if on_progress:
                    on_progress(i, len(doctor_names), name, error=msg)
                rows = []
                break
        per_doctor[name] = len(rows)
        all_rows.extend(rows)
        if on_progress:
            on_progress(i, len(doctor_names), name, count=len(rows))

    if output_path is None:
        output_path = _default_practice_table_path(doctor_names)
    out_path = Path(output_path)
    _write_xlsx(all_rows, out_path, sheet_name=sheet_name)
    return len(all_rows), out_path, per_doctor


def fetch_practice_table_parallel(
    doctor_names: list[str],
    reg_cfg: dict | None = None,
    config_path: Path | str | None = None,
    output_path: Path | str | None = None,
    sheet_name: str = "明细",
    on_progress=None,
    list_cache: dict[int, list[dict[str, str]]] | None = None,
    workers: int | None = None,
) -> tuple[int, Path, dict[str, int]]:
    """全量/大批量场景：列表只拉1次 → 并行预取全部 detail+muti → 纯内存组装。

    与 ``fetch_practice_table`` 输出口径一致，仅速度更快：
    - 跨医生共享 detail / muti 缓存（线程安全）
    - 姓名索引避免 N×5 次全表扫描
    - 并发预取 detail / muti，workers 默认按 CPU 推算
    返回 (总行数, 输出路径, 每医生行数)。
    """
    from src.minke_reg.soap import check_minke_service_route

    reg = reg_cfg or _load_reg_cfg(config_path)
    # 全量预取前硬拦截 fake-ip，避免 detail×上万次空转超时
    check_minke_service_route(str(reg.get("docUnitServiceUrl") or ""), on_fake_ip="error")
    session = login_minke_reg(reg)
    w = workers or default_workers(reg)

    if list_cache is None:
        list_cache = {}
    if on_progress:
        on_progress(0, len(doctor_names), "预拉全量列表")
    for st, _label, _kind in SEARCH_SPECS:
        if st not in list_cache:
            _fetch_list_for_other(reg, session, st, list_cache)

    # 姓名索引
    name_index = _build_name_index(list_cache)

    # 收集全部预取 key
    detail_keys, muti_reg_ids = _collect_prefetch_keys(name_index, doctor_names)
    if on_progress:
        on_progress(0, len(doctor_names), f"预取 detail×{len(detail_keys)} muti×{len(muti_reg_ids)} (workers={w})")

    # 并行预取 detail（内部会话失效自动重登录）
    if on_progress:
        on_progress(0, len(doctor_names), f"预取 detail {len(detail_keys)} 次")
    detail_holder = _prefetch_details(reg, session, detail_keys, w, on_progress)
    session = detail_holder["session"] if detail_holder else session

    # 并行预取 muti（内部会话失效自动重登录）
    if on_progress:
        on_progress(0, len(doctor_names), f"预取 muti {len(muti_reg_ids)} 次")
    muti_holder = _prefetch_muti(reg, session, muti_reg_ids, w, on_progress)
    session = muti_holder["session"] if muti_holder else session

    # 纯内存组装（无网络）
    all_rows: list[dict[str, str]] = []
    per_doctor: dict[str, int] = {}
    for i, name in enumerate(doctor_names, 1):
        rows = _assemble_rows_for_doctor(
            reg, session, name, name_index, _DETAIL_CACHE, _MUTI_REG_CACHE
        )
        per_doctor[name] = len(rows)
        all_rows.extend(rows)
        if on_progress:
            on_progress(i, len(doctor_names), name, count=len(rows))

    if output_path is None:
        output_path = _default_practice_table_path(doctor_names)
    out_path = Path(output_path)
    _write_xlsx(all_rows, out_path, sheet_name=sheet_name)
    return len(all_rows), out_path, per_doctor


# ---------- 电子证照独立流程 ----------

ELEC_HEADERS: tuple[str, ...] = (
    "姓名",
    "执业证书编码",
    "身份证号",
    "医师类别",
    "医师级别",
    "执业范围",
    "查看电子证照",
    "是否已申领电子证照",
)


def parse_elec_doctor_query(raw: str) -> tuple[str, str | None]:
    """解析 ``姓名`` 或 ``姓名:执业证书编号`` / ``姓名：执业证书编号``。"""
    text = (raw or "").strip()
    if not text:
        return "", None
    for sep in (":", "："):
        if sep in text:
            name, licence = text.split(sep, 1)
            name = name.strip()
            licence = licence.strip()
            return name, licence or None
    return text, None


def _normalize_licence(code: str | None) -> str:
    return (code or "").strip().upper()


def _iter_name_matches(
    list_cache: dict[int, list[dict[str, str]]], name: str,
) -> list[dict[str, str]]:
    """st=1 优先，再 st=8；同名单内保持接口返回顺序。"""
    matched: list[dict[str, str]] = []
    for st in (1, 8):
        for r in list_cache.get(st) or []:
            if (r.get("Doctor_Name") or "").strip() == name:
                matched.append(r)
    return matched


def _licence_on_row(
    reg: dict,
    session,
    row: dict[str, str],
    *,
    fetch_detail_if_missing: bool,
) -> str:
    """读行上的执业证书编码；名单无该字段时可补一次详情。"""
    code = (row.get("WorkLicenceCode") or "").strip()
    if code:
        return code
    if not fetch_detail_if_missing:
        return ""
    detail = _detail_map(
        reg,
        session,
        str(row.get("Doctor_GID", "")),
        str(row.get("Doctor_RegisterGID", "")),
    )
    code = (detail.get("WorkLicenceCode") or "").strip()
    if code:
        row["WorkLicenceCode"] = code
    return code


def _find_reg_row(
    list_cache: dict[int, list[dict[str, str]]],
    name: str,
    *,
    work_licence: str | None = None,
    reg: dict | None = None,
    session=None,
) -> dict[str, str] | None:
    """按姓名定位注册行；提供执业证书编号时按「姓名+证号」唯一定位。

    - 未给证号：兼容旧行为，取 st=1/8 中该姓名的首条
    - 给了证号：在同名候选人中匹配 WorkLicenceCode；名单无证号时补拉详情
    """
    candidates = _iter_name_matches(list_cache, name)
    if not candidates:
        return None

    want = _normalize_licence(work_licence)
    if not want:
        return candidates[0]

    if reg is None or session is None:
        raise ValueError("按执业证书编号匹配时需要 reg/session")

    for r in candidates:
        code = _licence_on_row(reg, session, r, fetch_detail_if_missing=True)
        if _normalize_licence(code) == want:
            return r
    return None


def fetch_elec_license(
    doctor_names: list[str],
    reg_cfg: dict | None = None,
    config_path: Path | str | None = None,
    output_path: Path | str | None = None,
    on_progress=None,
) -> tuple[int, Path]:
    """获取多名医生的电子证照预览 URL + 申领状态，写入 Excel。

    ``doctor_names`` 每项可为 ``姓名`` 或 ``姓名:执业证书编号``。
    带证号时按姓名+证号唯一定位；仅姓名时取该名在 st=1/8 的首条（兼容旧用法）。
    """
    queries = [parse_elec_doctor_query(raw) for raw in doctor_names]
    queries = [(n, lic) for n, lic in queries if n]
    if not queries:
        raise ValueError("请指定医生姓名（可选 姓名:执业证书编号）")

    reg = reg_cfg or _load_reg_cfg(config_path)
    session = login_minke_reg(reg)

    list_cache: dict[int, list[dict[str, str]]] = {}
    if on_progress:
        on_progress(0, len(queries), "预拉列表(st=1/8)")
    for st in (1, 8):
        _fetch_list_for_other(reg, session, st, list_cache)

    rows_out: list[dict[str, str]] = []
    for i, (name, licence) in enumerate(queries, 1):
        label = f"{name}:{licence}" if licence else name
        if on_progress:
            on_progress(i, len(queries), label)
        row = _find_reg_row(
            list_cache,
            name,
            work_licence=licence,
            reg=reg,
            session=session,
        )
        if not row:
            miss = "未找到该医生（姓名+执业证书编号无匹配）" if licence else "未找到该医生"
            rows_out.append(
                {
                    "姓名": name,
                    "执业证书编码": licence or "",
                    "查看电子证照": "",
                    "是否已申领电子证照": miss,
                }
            )
            if on_progress:
                on_progress(i, len(queries), label, count=0, error="未找到")
            continue
        doctor_gid = str(row.get("Doctor_GID", ""))
        register_gid = str(row.get("Doctor_RegisterGID", ""))
        resolved_licence = _licence_on_row(
            reg, session, row, fetch_detail_if_missing=bool(licence)
        ) or (licence or "")
        elec_url = make_electronic_license_url(doctor_gid, register_gid)
        applied = check_elec_applied(elec_url)
        rows_out.append(
            {
                "姓名": name,
                "执业证书编码": resolved_licence,
                "身份证号": row.get("IDCard", ""),
                "医师类别": row.get("Doctor_SortName", ""),
                "医师级别": row.get("Doctor_LevelName", ""),
                "执业范围": row.get("Subject_Name", ""),
                "查看电子证照": elec_url,
                "是否已申领电子证照": applied,
            }
        )
        if on_progress:
            on_progress(i, len(queries), label, count=1, error=applied or None)

    if output_path is None:
        output_path = Path("workspace/artifacts") / (
            f"elec_license_batch{len(queries)}.xlsx"
            if len(queries) > 1
            else f"elec_license_{queries[0][0]}.xlsx"
        )
    out_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "电子证照"[:31]
    bold = Font(bold=True)
    for col_idx, h in enumerate(ELEC_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows_out:
        ws.append([r.get(h, "") for h in ELEC_HEADERS])
    widths = {
        "姓名": 8,
        "执业证书编码": 22,
        "身份证号": 20,
        "医师类别": 8,
        "医师级别": 10,
        "执业范围": 12,
        "查看电子证照": 60,
        "是否已申领电子证照": 22,
    }
    for col_idx, h in enumerate(ELEC_HEADERS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(h, 12)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(ELEC_HEADERS))
    last_row = max(1, len(rows_out) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(rows_out), out_path
