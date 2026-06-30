from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from src.reconcile.field_mapping import FIELD_LABELS
from src.reconcile.to_supplement import capture_meta, is_create_op, iter_payloads, supplement_fields

_ROSTER_HEADERS = ["name", "certCode", "idCard"]
_ROSTER_LABELS = ["姓名", "执业证书编号", "身份证"]

_SHEET_LIANOU_HAS = "莲藕有机构端无"
_SHEET_EXPORT_HAS = "机构端有莲藕无"
_SHEET_SUPPLEMENT = "需补充名单"

_SUPPLEMENT_HEADERS = [
    "name",
    "certCode",
    "idCard",
    "listEntry",
    "operation",
    "hospital",
    "aId",
    "supplementFields",
]
_SUPPLEMENT_LABELS = [
    "姓名",
    "执业证书编号",
    "身份证",
    "执业列表",
    "操作",
    "医院",
    "aId",
    "需补充字段",
]

_LIST_ENTRY_LABELS = {
    "Main": "主执业",
    "Multi": "多执业",
}

_IMAGE_FIELDS = frozenset({"institutionBase", "healthCommissionBase"})

# 与 to_supplement._SUPPLEMENT_KEYS 顺序一致，便于阅读
_SUPPLEMENT_FIELD_ORDER = (
    "medicalInstitutionType",
    "practiceProvince",
    "practiceCity",
    "hospital",
    "hospitalLevel",
    "professionalList",
    "recordDate",
    "recordExpireDate",
    "institutionBase",
    "healthCommissionBase",
)


def _write_roster_sheet(ws, records: list[dict[str, Any]]) -> None:
    header_font = Font(bold=True)
    for col_idx, label in enumerate(_ROSTER_LABELS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, key in enumerate(_ROSTER_HEADERS, start=1):
            value = record.get(key, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"


def _list_entry_label(value: str | None) -> str:
    key = str(value or "Main").strip()
    return _LIST_ENTRY_LABELS.get(key, key)


def _operation_label(payload: dict[str, Any]) -> str:
    return "新增" if is_create_op(payload) else "更新"


def _format_supplement_fields(fields: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in _SUPPLEMENT_FIELD_ORDER:
        if key not in fields:
            continue
        label = FIELD_LABELS.get(key, key)
        value = fields.get(key)
        if key in _IMAGE_FIELDS and (value is None or value == ""):
            parts.append(f"{label}(待采图)")
        else:
            parts.append(label)
    return ";".join(parts)


def build_supplement_report_rows(
    to_submit: list[dict[str, Any]] | dict[str, Any],
) -> list[dict[str, Any]]:
    """将 to_submit 操作体展开为需补充名单行。"""
    rows: list[dict[str, Any]] = []
    for payload in iter_payloads(to_submit):
        meta = capture_meta(payload)
        fields = supplement_fields(payload)
        id_card = payload.get("iDCard") or meta.get("idCard") or ""
        rows.append(
            {
                "name": payload.get("doctorName") or "",
                "certCode": meta.get("certCode") or payload.get("practicingCertCode") or "",
                "idCard": id_card,
                "listEntry": _list_entry_label(meta.get("listEntry")),
                "operation": _operation_label(payload),
                "hospital": meta.get("hospital") or "",
                "aId": payload.get("aId") or payload.get("AId") or "",
                "supplementFields": _format_supplement_fields(fields),
            }
        )
    return rows


def _write_supplement_sheet(ws, records: list[dict[str, Any]]) -> None:
    header_font = Font(bold=True)
    for col_idx, label in enumerate(_SUPPLEMENT_LABELS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, key in enumerate(_SUPPLEMENT_HEADERS, start=1):
            value = record.get(key, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"


def save_reconcile_report(
    *,
    lianou_only: list[dict[str, Any]],
    export_only: list[dict[str, Any]],
    to_submit: list[dict[str, Any]] | None = None,
    output_path: Path,
) -> None:
    """写入单份核对报告 xlsx：未匹配名单 + 需补充名单。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_lianou = wb.create_sheet(title=_SHEET_LIANOU_HAS[:31])
    _write_roster_sheet(ws_lianou, lianou_only)

    ws_export = wb.create_sheet(title=_SHEET_EXPORT_HAS[:31])
    _write_roster_sheet(ws_export, export_only)

    supplement_rows = build_supplement_report_rows(to_submit or [])
    ws_supplement = wb.create_sheet(title=_SHEET_SUPPLEMENT[:31])
    _write_supplement_sheet(ws_supplement, supplement_rows)

    wb.save(output_path)


def save_missing_roster(
    records: list[dict[str, Any]],
    output_path: Path,
    *,
    title: str = "未匹配名单",
) -> None:
    """单 sheet 名单（保留供特殊导出场景）。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _write_roster_sheet(ws, records)
    wb.save(output_path)
