from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import project_root

COLUMN_DEFS: list[tuple[str, str, str]] = [
    ("aId", "主键ID", "aId"),
    ("doctorFileId", "医生档案ID", "doctorFileId"),
    ("doctorName", "医生姓名", "doctorName"),
    ("hospital", "医院名称", "hospital"),
    ("medicalInstitutionType", "医疗机构类型", "medicalInstitutionType"),
    ("medicalInstitutionTypeLabel", "医疗机构类型说明", "_medicalInstitutionTypeLabel"),
    ("practiceProvince", "省份", "practiceProvince"),
    ("practiceCity", "城市", "practiceCity"),
    ("hospitalLevel", "医院等级", "hospitalLevel"),
    ("hospitalLevelLabel", "医院等级说明", "_hospitalLevelLabel"),
    ("professionalList", "执业范围", "_professionalListLabel"),
    ("recordDate", "备案日期", "recordDate"),
    ("recordExpireDate", "备案到期日期", "recordExpireDate"),
    ("healthCommissionBase", "卫健委图片", "healthCommissionBase"),
    ("institutionBase", "机构端图片", "institutionBase"),
    ("updateField", "缺失字段", "updateField"),
]


def autosize_columns(ws) -> None:
    for col_idx, (_, header, _) in enumerate(COLUMN_DEFS, start=1):
        max_len = len(header)
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_len + 2, 60)


def save_xlsx(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "医生医疗机构信息"

    header_font = Font(bold=True)
    for col_idx, (_, header, _) in enumerate(COLUMN_DEFS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, (_, _, source_key) in enumerate(COLUMN_DEFS, start=1):
            value = record.get(source_key, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    wb.save(output_path)


def default_output_path(api_cfg: dict[str, Any]) -> Path:
    output_dir = Path(api_cfg.get("outputDir") or "exports")
    if not output_dir.is_absolute():
        output_dir = project_root() / output_dir
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return output_dir / f"医生医疗机构信息-{timestamp}.xlsx"
